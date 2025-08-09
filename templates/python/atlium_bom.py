import logging
import os
import subprocess
import sys
import tkinter as tk
from collections.abc import Sequence
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, font, messagebox
from typing import Any, TypeGuard

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("bom_sorter")

# Material design color scheme
MATERIAL_COLORS = {
    "background": "#121212",  # Dark background
    "surface": "#1E1E1E",  # Slightly lighter surface
    "primary": "#BB86FC",  # Purple
    "primary_variant": "#3700B3",
    "secondary": "#03DAC6",  # Teal
    "on_background": "#FFFFFF",  # White text
    "on_surface": "#E1E1E1",  # Light grey text
    "on_primary": "#000000",  # Black text on primary color
    "on_secondary": "#000000",  # Black text on secondary color
    "error": "#CF6679",  # Error color
}

def categorize_order(footprint: str) -> int:
    """
    Categorize components based on Footprint prefix for sorting purposes.

    Args:
        footprint: The footprint string to categorize

    Returns:
        Integer representing sort priority (lower = higher priority)
    """
    desired_order = ["CAP_", "RES_", "FB_", "IND_"]

    for i, prefix in enumerate(desired_order):
        if footprint.startswith(prefix):
            return i

    # Special case for connectors and similar components
    if "SOLDER" in footprint or footprint.startswith("JST_") or "PicoBlade" in footprint:
        return len(desired_order) + 1

    # Default for other uncategorized components
    return len(desired_order)


def process_file(raw_file_path: str, output_file_path: str) -> str:
    """
    Process the BOM file, sort it by component types, and apply formatting.

    Args:
        raw_file_path: Path to the input Excel file
        output_file_path: Path where the processed file will be saved

    Returns:
        Path to the processed file
    """
    try:
        logger.info(f"Reading Excel file: {raw_file_path}")
        df = pd.read_excel(raw_file_path)
        logger.info(f"Columns found: {df.columns.tolist()}")

        # Define the desired column order
        desired_column_order = ["Quantity", "Designator", "Comment", "Description", "Footprint", "LibRef", "Layer"]

        # Filter to include only columns present in the DataFrame
        existing_columns = [col for col in desired_column_order if col in df.columns]
        logger.info(f"Existing columns used for sorting: {existing_columns}")

        # Reorder the columns
        df = df[existing_columns]

        # Add sorting helper columns
        df["Footprint_Order"] = df["Footprint"].apply(categorize_order)
        df["Layer_Order"] = df["Layer"].map({"Top": 0, "Bottom": 1})
        logger.info("Applied categorization functions")

        # Sort by Layer first, then by component type
        sorted_df = df.sort_values(by=["Layer_Order", "Footprint_Order", "Footprint"])

        # Remove helper columns
        sorted_df = sorted_df.drop(columns=["Footprint_Order", "Layer_Order"])

        # Save the sorted DataFrame to Excel
        sorted_df.to_excel(output_file_path, index=False)
        logger.info(f"Saved sorted DataFrame to {output_file_path}")

        # Apply visual styling
        apply_styles(output_file_path)

        return output_file_path

    except Exception as e:
        logger.error(f"Error processing file: {e}", exc_info=True)
        raise


def apply_styles(output_file_path: str) -> None:
    """
    Apply color-coding and formatting to the Excel file based on component types.

    Args:
        output_file_path: Path to the Excel file to format
    """
    try:
        logger.info(f"Applying styles to {output_file_path}")
        wb = load_workbook(output_file_path)
        ws = wb.active

        # Define color mappings for different components
        color_mapping = {
            "CAP_": {
                "0402": "FFFFE0",  # Light yellow
                "0603": "FFFACD",  # Lemon chiffon
                "0805": "FAFAD2",  # Light goldenrod yellow
                "1206": "FFEFD5",  # Papaya whip
                "1210": "FFD700",  # Gold
            },
            "RES_": {
                "0402": "FFC0C0",  # Light red
                "0603": "FFB6C1",  # Light pink
                "0805": "FF9999",  # Light coral
                "1206": "FF6347",  # Tomato
                "1210": "FF4500",  # Orange red
            },
            "L_FB": "FFCCCC",  # Pink
            "D": "ADD8E6",  # Light blue
        }

        # Find required column indices
        column_indices = find_column_indices(ws, ["Footprint", "Designator"])

        # Apply styles to each row
        apply_row_styling(ws, column_indices, color_mapping)

        # Adjust column widths for better readability
        adjust_column_widths(ws)

        wb.save(output_file_path)
        logger.info(f"Saved styled workbook to {output_file_path}")

    except Exception as e:
        logger.error(f"Error applying styles: {e}", exc_info=True)
        raise


def is_column_names(names: object) -> TypeGuard[Sequence[str]]:
    """Type guard to ensure column names are a sequence of strings."""
    return isinstance(names, Sequence) and all(isinstance(name, str) for name in names)


def find_column_indices(worksheet: Worksheet, column_names: Sequence[str]) -> dict[str, int]:
    """
    Find the indices of column names in a worksheet.

    Args:
        worksheet: The worksheet to search in
        column_names: List of column names to find

    Returns:
        Dictionary mapping column names to their indices
    """
    if not is_column_names(column_names):
        raise TypeError("column_names must be a sequence of strings")

    indices: dict[str, int] = {}

    for col in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=1, column=col).value
        if isinstance(header, str) and header in column_names:
            indices[header] = col

    # Verify all required columns were found
    missing = [name for name in column_names if name not in indices]
    if missing:
        raise ValueError(f"Required columns not found: {', '.join(missing)}")

    return indices


def apply_row_styling(worksheet: Worksheet, column_indices: dict[str, int], color_mapping: dict[str, Any]) -> None:
    """
    Apply color styling to rows based on component types.

    Args:
        worksheet: The worksheet to style
        column_indices: Dictionary mapping column names to indices
        color_mapping: Color definitions for component types
    """
    footprint_col = column_indices["Footprint"]
    designator_col = column_indices["Designator"]

    # Apply styles to each row
    for row in worksheet.iter_rows(min_row=2):  # Skip header row
        footprint = row[footprint_col - 1].value
        designator = row[designator_col - 1].value

        # Skip rows without valid data
        if not isinstance(footprint, str) or not isinstance(designator, str):
            continue

        # Determine fill color based on component type and apply it
        if fill := get_component_fill(footprint, designator, color_mapping):
            for cell in row:
                cell.fill = fill


def get_component_fill(footprint: str, designator: str, color_mapping: dict[str, Any]) -> PatternFill | None:
    """
    Determine the color fill for a component based on type and size.

    Args:
        footprint: Component footprint string
        designator: Component designator string
        color_mapping: Color definitions dictionary

    Returns:
        PatternFill object or None if no fill should be applied
    """
    # Use pattern matching for component type determination
    match designator, footprint:
        case _ if "L" in designator or "FB" in designator:
            return create_fill(color_mapping["L_FB"])

        case _ if "D" in designator:
            return create_fill(color_mapping["D"])

        case _ if footprint.startswith("CAP_"):
            size = next((s for s in color_mapping["CAP_"] if s in footprint), "0402")
            return create_fill(color_mapping["CAP_"][size])

        case _ if footprint.startswith("RES_"):
            size = next((s for s in color_mapping["RES_"] if s in footprint), "0402")
            return create_fill(color_mapping["RES_"][size])

        case _:
            return None


def create_fill(color: str) -> PatternFill:
    """Create a solid fill pattern with the given color."""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def adjust_column_widths(worksheet: Worksheet) -> None:
    """
    Adjust column widths based on content.

    Args:
        worksheet: The worksheet to adjust
    """
    for col in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        for cell in worksheet[column_letter]:
            try:
                if cell.value is not None:
                    # More efficient string length calculation
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue

        # Set column width with some padding
        worksheet.column_dimensions[column_letter].width = max_length + 2


def center_window(window: tk.Tk, width: int, height: int) -> None:
    """
    Center a tkinter window on the screen.

    Args:
        window: The tkinter window to center
        width: The width of the window
        height: The height of the window
    """
    # Get screen width and height
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    # Calculate position coordinates
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2

    # Set the window position
    window.geometry(f"{width}x{height}+{x}+{y}")


class HoverButton(tk.Button):
    """Button widget with hover effect for material design."""

    def __init__(self, master=None, **kwargs):
        self.background = kwargs.pop('background', MATERIAL_COLORS["primary"])
        self.hover_background = kwargs.pop('hover_background', MATERIAL_COLORS["primary_variant"])
        self.foreground = kwargs.pop('foreground', MATERIAL_COLORS["on_primary"])
        self.hover_foreground = kwargs.pop('hover_foreground', MATERIAL_COLORS["on_primary"])

        kwargs['background'] = self.background
        kwargs['foreground'] = self.foreground
        kwargs['borderwidth'] = kwargs.get('borderwidth', 0)
        kwargs['relief'] = kwargs.get('relief', 'flat')

        super().__init__(master, **kwargs)

        # Bind events for hover effect
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)

    def _on_enter(self, event):
        """Change appearance on mouse enter."""
        self.config(background=self.hover_background, foreground=self.hover_foreground)

    def _on_leave(self, event):
        """Restore original appearance on mouse leave."""
        self.config(background=self.background, foreground=self.foreground)


def browse_and_sort() -> None:
    """
    Handle file selection and processing flow.
    Uses the input directory as the output directory
    and only asks once for file location.
    """
    # Open file dialog to select the input BOM file
    raw_file_path = filedialog.askopenfilename(title="Select BOM file to process", initialdir=os.getcwd(), filetypes=[("Excel files", "*.xlsx")])

    if not raw_file_path:
        return  # User canceled

    # Get file info for output
    input_path = Path(raw_file_path)
    file_dir = input_path.parent
    original_name = input_path.stem

    # Generate timestamp for the output filename
    timestamp = datetime.now().strftime("%y%m%d@%H%M%S")
    output_filename = f"{original_name}_sorted_{timestamp}.xlsx"

    # Create the output path in the same directory as the input
    output_file_path = file_dir / output_filename

    try:
        # Process the file
        processed_file = process_file(raw_file_path, str(output_file_path))

        messagebox.showinfo("Success", f"File processed successfully and saved as:\n{output_filename}")

        # Ask if the user wants to open the output file
        if messagebox.askyesno("Open File", "Do you want to open the processed file in Excel?"):
            open_file(processed_file)

    except Exception as e:
        logger.error(f"Error in browse_and_sort: {e}", exc_info=True)
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def open_file(file_path: str) -> None:
    """
    Open a file with the system's default application.

    Args:
        file_path: Path to the file to open
    """
    try:
        match os.name:
            case "nt":  # Windows
                os.startfile(file_path)
            case "posix":  # macOS or Linux
                subprocess.call(["open", file_path] if sys.platform == "darwin" else ["xdg-open", file_path])
            case _:
                logger.warning(f"Unsupported OS: {os.name}")
    except Exception as e:
        logger.error(f"Could not open file: {e}")
        messagebox.showerror("Error", f"Could not open the file: {str(e)}")


def create_material_theme(root: tk.Tk) -> None:
    """
    Apply material design theme to the root window.

    Args:
        root: The root tkinter window
    """
    # Configure the root window
    root.configure(bg=MATERIAL_COLORS["background"])

    # Configure the default font
    default_font = font.Font(family="Segoe UI", size=10)
    root.option_add("*Font", default_font)

    # Configure the message dialogs
    root.option_add("*Dialog.msg.font", default_font)

    # Configure button appearance
    root.option_add("*Button.Background", MATERIAL_COLORS["primary"])
    root.option_add("*Button.Foreground", MATERIAL_COLORS["on_primary"])
    root.option_add("*Button.Relief", "flat")
    root.option_add("*Button.BorderWidth", 0)

    # Configure label appearance
    root.option_add("*Label.Background", MATERIAL_COLORS["background"])
    root.option_add("*Label.Foreground", MATERIAL_COLORS["on_background"])


def main() -> None:
    """Initialize and run the application."""
    # Set up the main application window
    root = tk.Tk()
    root.title("BOM Sorter")

    # Apply material design theme
    create_material_theme(root)

    # Set window dimensions and center it
    window_width, window_height = 400, 180
    center_window(root, window_width, window_height)

    # Create a frame with padding
    main_frame = tk.Frame(root, bg=MATERIAL_COLORS["background"], padx=20, pady=20)
    main_frame.pack(fill="both", expand=True)

    # Create fonts
    title_font = font.Font(family="Segoe UI", size=16, weight="bold")
    normal_font = font.Font(family="Segoe UI", size=10)
    button_font = font.Font(family="Segoe UI", size=11)
    small_font = font.Font(family="Segoe UI", size=8)

    # Add title label
    title_label = tk.Label(
        main_frame,
        text="Altium BOM Sorter",
        font=title_font,
        bg=MATERIAL_COLORS["background"],
        fg=MATERIAL_COLORS["on_background"],
    )
    title_label.pack(pady=(0, 10))

    # Add description label
    description_label = tk.Label(
        main_frame,
        text="Process and format Altium BOM files",
        bg=MATERIAL_COLORS["background"],
        fg=MATERIAL_COLORS["on_surface"],
        font=normal_font,
    )
    description_label.pack(pady=(0, 20))

    # Main action button
    process_button = HoverButton(
        main_frame,
        text="Select and Process BOM File",
        command=browse_and_sort,
        font=button_font,
        padx=15,
        pady=8,
        cursor="hand2",  # Hand cursor on hover
    )
    process_button.pack()

    # Version label
    version_label = tk.Label(
        main_frame,
        text="v1.0.0",
        bg=MATERIAL_COLORS["background"],
        fg=MATERIAL_COLORS["on_surface"],
        font=small_font,
    )
    version_label.pack(side="bottom", pady=(20, 0))

    # Start the GUI event loop
    root.mainloop()


if __name__ == "__main__":
    main()
